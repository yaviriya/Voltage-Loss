import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, time, date, timedelta
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict

class VoltageAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Voltage Loss Calculation")
        self.root.geometry("600x600")
        
        # ตัวแปรเก็บพาธของไฟล์
        self.input_paths = [None, None, None]
        self.multiply_factor_var = tk.StringVar(value="30")  # ตั้งค่าเริ่มต้นเป็น 30
        
        # ตัวแปรสำหรับเก็บผลรวมของ P Loss แยกตามสี
        self.red_total_p_loss = 0
        self.green_total_p_loss = 0
        self.blue_total_p_loss = 0
        
        # เตรียมสีที่จะใช้
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')   # สีแดง
        self.green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid') # สีเขียว
        self.blue_fill = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')  # สีฟ้า
        
        # วันหยุดราชการ
        self.holidays = {
            date(2025, 1, 1),
            date(2025, 2, 12),
            date(2025, 4, 7),
            date(2025, 4, 14),
            date(2025, 4, 15),
            date(2025, 4, 16),
            date(2025, 5, 1),
            date(2025, 5, 5),
            date(2025, 5, 9),
            date(2025, 5, 12),
            date(2025, 6, 2),
            date(2025, 6, 3),
            date(2025, 7, 10),
            date(2025, 7, 11),
            date(2025, 7, 28),
            date(2025, 8, 11),
            date(2025, 8, 12),
            date(2025, 10, 13),
            date(2025, 10, 23),
            date(2025, 12, 5),
            date(2025, 12, 10),
            date(2025, 12, 31)
            # ใส่วันอื่น ๆ ตามต้องการ
        }
        
        self.create_widgets()
    
    def create_widgets(self):
        """สร้าง UI elements"""
        # สร้างเฟรมสำหรับเลือกไฟล์
        file_frame = ttk.LabelFrame(self.root, text="Upload Excel Files")
        file_frame.pack(fill="x", padx=10, pady=10)
        
        # สร้างปุ่มและแสดงพาธของไฟล์ที่เลือก
        file_labels = ["Voltage", "Current", "Power Factor"]
        self.file_path_vars = [tk.StringVar() for _ in range(3)]
        
        for i, label in enumerate(file_labels):
            ttk.Label(file_frame, text=label).grid(row=i, column=0, sticky="w", padx=5, pady=5)
            ttk.Entry(file_frame, textvariable=self.file_path_vars[i], width=50).grid(row=i, column=1, padx=5, pady=5)
            ttk.Button(file_frame, text="Browse...", command=lambda idx=i: self.browse_file(idx)).grid(row=i, column=2, padx=5, pady=5)
        
        # สร้างเฟรมสำหรับตั้งค่าพารามิเตอร์
        param_frame = ttk.LabelFrame(self.root, text="CT")
        param_frame.pack(fill="x", padx=10, pady=10)
        
        # เพิ่มช่องสำหรับกรอกค่า CT Ratio (multiply_factor)
        ttk.Label(param_frame, text="Ratio:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(param_frame, textvariable=self.multiply_factor_var, width=10).grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        # สร้างเฟรมสำหรับเลือกเฟส
        phase_frame = ttk.LabelFrame(self.root, text="Select Phase")
        phase_frame.pack(fill="x", padx=10, pady=10)
        
        # สร้าง Radio Buttons สำหรับเลือกเฟส
        self.phase_var = tk.StringVar(value="B")  # ค่าเริ่มต้นคือเฟส B
        ttk.Radiobutton(phase_frame, text="Phase A", variable=self.phase_var, value="A").pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(phase_frame, text="Phase B", variable=self.phase_var, value="B").pack(side=tk.LEFT, padx=20)
        ttk.Radiobutton(phase_frame, text="Phase C", variable=self.phase_var, value="C").pack(side=tk.LEFT, padx=20)
        
        # สร้างปุ่มเริ่มการคำนวณ
        ttk.Button(self.root, text="Execute", command=self.process_files).pack(pady=20)
        
        # สร้างเฟรมสำหรับแสดงผลลัพธ์
        result_frame = ttk.LabelFrame(self.root, text="Results")
        result_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # สร้างตัวแปรสำหรับแสดงผลลัพธ์
        self.output_path_var = tk.StringVar()
        self.red_p_loss_var = tk.StringVar()
        self.green_p_loss_var = tk.StringVar()
        self.blue_p_loss_var = tk.StringVar()
        self.total_p_loss_var = tk.StringVar()
        
        # แสดงพาธของไฟล์ผลลัพธ์
        ttk.Label(result_frame, text="Output File:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.output_path_var, width=70, state="readonly").grid(row=0, column=1, padx=5, pady=5)
        
        # แสดงผลรวมของ P Loss แยกตามสี
        ttk.Label(result_frame, text="P Loss Peak:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        entry_red = ttk.Entry(result_frame, textvariable=self.red_p_loss_var, width=15, state="readonly")
        entry_red.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=1, column=1, sticky="w", padx=(120, 5), pady=5)
        
        ttk.Label(result_frame, text="P Loss Off-Peak:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        entry_green = ttk.Entry(result_frame, textvariable=self.green_p_loss_var, width=15, state="readonly")
        entry_green.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=2, column=1, sticky="w", padx=(120, 5), pady=5)
        
        ttk.Label(result_frame, text="P Loss Holiday:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        entry_blue = ttk.Entry(result_frame, textvariable=self.blue_p_loss_var, width=15, state="readonly")
        entry_blue.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=3, column=1, sticky="w", padx=(120, 5), pady=5)
        
        ttk.Label(result_frame, text="Total P Loss:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        entry_total = ttk.Entry(result_frame, textvariable=self.total_p_loss_var, width=15, state="readonly")
        entry_total.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=4, column=1, sticky="w", padx=(120, 5), pady=5)
        
        # สร้างปุ่มออกจากโปรแกรม
        ttk.Button(self.root, text="Exit", command=self.root.destroy).pack(pady=10)
    
    def browse_file(self, index):
        """เปิดหน้าต่างให้เลือกไฟล์ Excel"""
        file_path = filedialog.askopenfilename(
            title=f"Select Excel File {index+1}",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_paths[index] = file_path
            self.file_path_vars[index].set(file_path)
    
    def is_weekend_or_holiday(self, dt):
        """ตรวจสอบว่าเป็นวันเสาร์-อาทิตย์ หรือวันหยุดราชการ"""
        # ถ้าเป็นวันหยุดราชการ (วันแบบ date ตรงกับในเซต holidays) ก็จะถือเป็นวันหยุด
        if dt.date() in self.holidays:
            return True
        # ตรวจสอบวันเสาร์-อาทิตย์
        # Monday=0, Tuesday=1, ... Sunday=6
        if dt.weekday() >= 5:  # 5=Saturday, 6=Sunday
            return True
        return False
    
    def find_header_row(self, worksheet):
        """หาแถวที่เป็นหัวตาราง โดยหาแถวแรกที่คอลัมน์ A มีรูปแบบวันที่"""
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
            cell_a_value = str(row[0].value).strip() if row[0].value else ""
            # ตรวจสอบว่าแถวนี้มีรูปแบบวันที่หรือไม่
            if self.is_valid_datetime(cell_a_value):
                # ถ้าเจอแถวที่มีรูปแบบวันที่ ให้ถือว่าแถวก่อนหน้าเป็นหัวตาราง
                # แต่ต้องตรวจสอบว่าไม่ใช่แถวแรก
                if row_idx > 1:
                    return row_idx - 1
                else:
                    return row_idx  # กรณีเป็นแถวแรก ให้ถือว่าแถวนั้นเป็นหัวตาราง
        return None  # ไม่พบแถวที่มีรูปแบบวันที่
    
    def is_valid_datetime(self, text):
        """ตรวจสอบว่าข้อความเป็นวันที่หรือไม่"""
        if not text:
            return False
        # ตรวจสอบว่าข้อความมีรูปแบบตัวเลขและเครื่องหมาย / หรือ - หรือไม่
        return any(c.isdigit() for c in text) and ('/' in text or '-' in text)
    
    def read_excel_data(self, file_path):
        """อ่านข้อมูลจากไฟล์ Excel"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        header_row = self.find_header_row(ws)
        
        if header_row is None:
            messagebox.showwarning("Warning", f"ไม่พบแถวหัวตารางในไฟล์ {os.path.basename(file_path)}")
            return None
        
        data = {}
        is_thai_date = False  # เริ่มต้นกำหนดให้ไม่ใช่วันที่ไทย
        
        # ตรวจสอบแถวแรกที่มีข้อมูลวันที่เพื่อดูว่าเป็นวันที่ไทยหรือไม่
        for row in ws.iter_rows(min_row=header_row + 1):
            raw_value = str(row[0].value) if row[0].value is not None else ""
            if self.is_valid_datetime(raw_value):
                clean_value = raw_value.replace('\xa0', '').strip()
                
                # แยกวันที่กับเวลา
                # ตรวจสอบหลายรูปแบบ: มีช่องว่างระหว่างวันที่กับเวลา หรือไม่มีช่องว่าง
                if ' ' in clean_value:
                    # รูปแบบ "วันที่ เวลา"
                    parts = clean_value.split()
                    date_str = parts[0]
                else:
                    # ถ้าไม่มีช่องว่าง ให้ตรวจสอบว่ามีรูปแบบวันที่หรือไม่
                    date_str = clean_value
                
                # ตรวจสอบรูปแบบวันที่ (dd/mm/yyyy หรือ dd-mm-yyyy)
                if '/' in date_str:
                    day, month, year = date_str.split('/')
                elif '-' in date_str:
                    day, month, year = date_str.split('-')
                else:
                    continue  # ข้ามถ้าไม่ใช่รูปแบบที่รองรับ
                
                try:
                    year_int = int(year)
                    # ถ้า year - 543 > 2000 แสดงว่าเป็นวันที่ไทย
                    if year_int - 543 > 2000:
                        is_thai_date = True
                except ValueError:
                    pass
                break  # ตรวจสอบเฉพาะแถวแรกที่มีวันที่
        
        # อ่านข้อมูลทั้งหมด
        for row in ws.iter_rows(min_row=header_row + 1):
            try:
                raw_value = str(row[0].value) if row[0].value is not None else ""
                if not self.is_valid_datetime(raw_value):
                    continue

                clean_value = raw_value.replace('\xa0', '').strip()
                
                # แยกวันที่กับเวลา
                if ' ' in clean_value:
                    parts = clean_value.split()
                    date_str = parts[0]
                    time_str = parts[1]
                else:
                    # กรณีที่ไม่มีช่องว่าง ต้องแยกวันที่กับเวลาเอง
                    if '/' in clean_value:
                        # หารูปแบบ dd/mm/yyyy
                        date_parts = clean_value.split('/')
                        if len(date_parts) >= 3:
                            date_str = '/'.join(date_parts[:3])
                            time_str = "00:00"  # ถ้าไม่มีเวลา ใช้เวลา 00:00
                        else:
                            continue
                    elif '-' in clean_value:
                        # หารูปแบบ dd-mm-yyyy
                        date_parts = clean_value.split('-')
                        if len(date_parts) >= 3:
                            date_str = '-'.join(date_parts[:3])
                            time_str = "00:00"  # ถ้าไม่มีเวลา ใช้เวลา 00:00
                        else:
                            continue
                    else:
                        continue
                
                # ตรวจสอบรูปแบบวันที่ (dd/mm/yyyy หรือ dd-mm-yyyy)
                if '/' in date_str:
                    day, month, year = date_str.split('/')
                    date_format = "%d/%m/%Y"
                elif '-' in date_str:
                    day, month, year = date_str.split('-')
                    date_format = "%d-%m-%Y"
                else:
                    continue  # ข้ามถ้าไม่ใช่รูปแบบที่รองรับ
                
                # ถ้าเป็นวันที่ไทย ให้ลบพ.ศ. 543
                if is_thai_date:
                    year = str(int(year) - 543)
                
                # แปลงรูปแบบวันที่ให้เป็น dd/mm/yyyy สำหรับการประมวลผลต่อไป
                std_date_str = f"{day}/{month}/{year}"
                
                # จัดการกับรูปแบบเวลา
                if ':' in time_str:
                    # รูปแบบ hh:mm:ss หรือ hh:mm
                    time_parts = time_str.split(':')
                    hour = time_parts[0]
                    minute = time_parts[1]
                    # ลบวินาทีออกถ้ามี
                    if len(minute) > 2:
                        minute = minute[:2]
                    
                    # เปลี่ยนรูปแบบเวลาให้เป็น hh.mm
                    std_time_str = f"{hour}.{minute}"
                else:
                    # รูปแบบอื่นๆ เช่น hh.mm
                    std_time_str = time_str
                
                # จัดการกับเวลา 24.00 หรือ 24:00
                if std_time_str == "24.00" or std_time_str == "24:00" or (hour == "24" if 'hour' in locals() else False):
                    # แปลงเป็นวันที่ถัดไป เวลา 00:00
                    dt_date = datetime.strptime(std_date_str, "%d/%m/%Y")
                    dt = dt_date + timedelta(days=1)
                else:
                    # กรณีเวลาปกติ
                    std_date_time_str = f"{std_date_str} {std_time_str}"
                    try:
                        dt = datetime.strptime(std_date_time_str, "%d/%m/%Y %H.%M")
                    except ValueError:
                        # หากแปลงไม่ได้ ลองเปลี่ยนรูปแบบเวลา
                        if '.' in std_time_str:
                            h, m = std_time_str.split('.')
                            std_time_str = f"{h}:{m}"
                        elif ':' in std_time_str:
                            h, m = std_time_str.split(':')
                            std_time_str = f"{h}.{m}"
                        std_date_time_str = f"{std_date_str} {std_time_str}"
                        try:
                            dt = datetime.strptime(std_date_time_str, "%d/%m/%Y %H.%M")
                        except ValueError:
                            dt = datetime.strptime(std_date_str, "%d/%m/%Y")  # ถ้าแปลงไม่ได้อีก ใช้แค่วันที่
                
                # เก็บข้อมูลทั้งแถว
                data[dt] = [cell.value for cell in row]
            except Exception as e:
                continue
        
        return data, header_row, ws
    
    def process_files(self):
        """ประมวลผลไฟล์ Excel ทั้งหมด"""
        # ตรวจสอบว่าได้เลือกไฟล์ครบทั้ง 3 ไฟล์หรือไม่
        if None in self.input_paths:
            messagebox.showerror("Error", "โปรดเลือกให้ครบทั้ง 3 ไฟล์เพื่อทำการคำนวณ")
            return
        
        try:
            # รับค่า multiply_factor จากช่องกรอก CT Ratio
            self.multiply_factor = float(self.multiply_factor_var.get())
        except ValueError:
            messagebox.showerror("Error", "กรุณาป้อนตัวคูณเป็นตัวเลข")
            return
        
        # รีเซ็ตตัวแปรผลรวม P Loss
        self.red_total_p_loss = 0
        self.green_total_p_loss = 0
        self.blue_total_p_loss = 0
        
        # อ่านข้อมูลจากทั้ง 3 ไฟล์
        all_data = []
        worksheets = []
        header_rows = []
        
        for path in self.input_paths:
            result = self.read_excel_data(path)
            if result:
                data, header_row, ws = result
                all_data.append(data)
                worksheets.append(ws)
                header_rows.append(header_row)
        
        # ตรวจสอบว่าอ่านข้อมูลได้ครบหรือไม่
        if len(all_data) != 3:
            messagebox.showerror("Error", "บางไฟล์ไม่สามารถอ่านได้ กรุณาตรวจสอบไฟล์ของคุณ")
            return
        
        # รวมข้อมูลจากทั้ง 3 ไฟล์
        merged_data = defaultdict(list)
        for dt in sorted(set().union(*[d.keys() for d in all_data])):
            row_data = []
            # เพิ่ม datetime เป็นคอลัมน์แรก
            # แปลงเวลา 00:00 ของวันถัดไปกลับเป็น 24.00 ของวันก่อนหน้า
            if dt.hour == 0 and dt.minute == 0:
                prev_day = dt - timedelta(days=1)
                datetime_str = prev_day.strftime("%d/%m/%Y 24.00")
            else:
                datetime_str = dt.strftime("%d/%m/%Y %H.%M")
            row_data.append(datetime_str)
            
            # เพิ่มข้อมูลจากแต่ละไฟล์ (ยกเว้นคอลัมน์ datetime)
            for data in all_data:
                if dt in data:
                    # ข้ามคอลัมน์แรก (datetime) และเพิ่มเฉพาะคอลัมน์ที่มีข้อมูล
                    row_data.extend([val for val in data[dt][1:] if val is not None])
                else:
                    # ถ้าไม่มีข้อมูลในไฟล์นั้น ให้ข้ามไป
                    continue
            merged_data[dt] = row_data
        
        # สร้างไฟล์ Excel ใหม่
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        
        # เขียนหัวตาราง
        headers = ["DateTime"]  # เพิ่มหัวตารางสำหรับ datetime
        
        # กำหนดชื่อหัวคอลัมน์ใหม่
        new_headers = [
            "V Phase A",
            "V Phase B",
            "V Phase C",
            "I Phase A",
            "I Phase B",
            "I Phase C",
            "Power Factor",
            "V Diff",
            "V loss",
            "P Loss"
        ]
        
        # เพิ่มชื่อหัวคอลัมน์ใหม่
        headers.extend(new_headers)
        
        # เขียนหัวตารางลงในไฟล์
        output_ws.append(headers)
        
        # หาตำแหน่งคอลัมน์ Power Factor (ไม่สนใจตัวอักษรใหญ่หรือเล็ก)
        power_factor_col = None
        for i, header in enumerate(headers, start=1):
            if header is not None and "power factor" in str(header).lower():
                power_factor_col = i
                break
        
        # หาตำแหน่งคอลัมน์ V Phase A, B, C
        v_phase_a_col = headers.index("V Phase A") + 1
        v_phase_b_col = headers.index("V Phase B") + 1
        v_phase_c_col = headers.index("V Phase C") + 1
        v_diff_col = headers.index("V Diff") + 1
        
        # ดึงเฟสที่เลือก
        phase = self.phase_var.get()
        
        # เขียนข้อมูลและใส่สี
        for dt, row_data in sorted(merged_data.items()):
            # แปลงค่า Power Factor ให้เป็นค่าบวก
            if power_factor_col is not None and len(row_data) >= power_factor_col:
                pf_value = row_data[power_factor_col - 1]
                if pf_value is not None:
                    try:
                        pf_float = float(pf_value)
                        if pf_float < 0:
                            row_data[power_factor_col - 1] = abs(pf_float)
                    except (ValueError, TypeError):
                        pass
            
            # คำนวณค่า V Diff
            try:
                # ตรวจสอบว่าข้อมูลมีเพียงพอหรือไม่
                max_index = max(v_phase_a_col - 1, v_phase_b_col - 1, v_phase_c_col - 1)
                if len(row_data) > max_index:
                    v_a = float(row_data[v_phase_a_col - 1]) if row_data[v_phase_a_col - 1] is not None else 0
                    v_b = float(row_data[v_phase_b_col - 1]) if row_data[v_phase_b_col - 1] is not None else 0
                    v_c = float(row_data[v_phase_c_col - 1]) if row_data[v_phase_c_col - 1] is not None else 0
                    
                    # คำนวณ V Diff ตามเฟสที่เลือก
                    if phase == 'A':
                        v_diff = ((v_b + v_c) / 2) - v_a
                    elif phase == 'B':
                        v_diff = ((v_a + v_c) / 2) - v_b
                    else:  # phase == 'C'
                        v_diff = ((v_a + v_b) / 2) - v_c
                else:
                    v_diff = None
                row_data.append(v_diff)  # เพิ่มค่า V Diff ต่อท้ายแถว
                
                # คำนวณค่า V loss ตามสูตร IF(V Diff>3,V Diff,0)
                if v_diff is not None and v_diff > 3:
                    v_loss = v_diff
                else:
                    v_loss = 0
                row_data.append(v_loss)  # เพิ่มค่า V loss ต่อท้ายแถว
                
                # คำนวณค่า P Loss ตามสูตร (V loss * I Phase X * Power Factor * multiply_factor) / 4000
                if power_factor_col is not None:
                    power_factor = float(row_data[power_factor_col - 1]) if row_data[power_factor_col - 1] is not None else 0
                    
                    # หาตำแหน่งคอลัมน์ I Phase ตามเฟสที่เลือก
                    if phase == 'A':
                        i_phase_col = headers.index("I Phase A") + 1
                    elif phase == 'B':
                        i_phase_col = headers.index("I Phase B") + 1
                    else:  # phase == 'C'
                        i_phase_col = headers.index("I Phase C") + 1
                    
                    if len(row_data) >= i_phase_col:
                        i_phase = float(row_data[i_phase_col - 1]) if row_data[i_phase_col - 1] is not None else 0
                        p_loss = (v_loss * i_phase * power_factor * self.multiply_factor) / 4000
                    else:
                        p_loss = 0
                else:
                    p_loss = 0
                row_data.append(p_loss)  # เพิ่มค่า P Loss ต่อท้ายแถว
            except (ValueError, TypeError, IndexError):
                row_data.append(None)  # V Diff เป็น None เมื่อคำนวณไม่ได้
                row_data.append(0)     # V loss เป็น 0 เมื่อคำนวณไม่ได้
                row_data.append(0)     # P Loss เป็น 0 เมื่อคำนวณไม่ได้
            
            output_ws.append(row_data)
            
            # หาตำแหน่งคอลัมน์ P Loss
            p_loss_col = headers.index("P Loss") + 1
            
            # ใส่สีในคอลัมน์ P Loss
            cell = output_ws.cell(row=output_ws.max_row, column=p_loss_col)
            
            # ตรวจสอบว่าเป็นเวลา 24.00 หรือไม่
            is_24_00 = dt.hour == 0 and dt.minute == 0
            
            if is_24_00:
                # ถ้าเป็นเวลา 24.00 ให้ใช้เงื่อนไขของวันก่อนหน้า
                prev_day = dt - timedelta(days=1)
                if self.is_weekend_or_holiday(prev_day):
                    cell.fill = self.blue_fill
                    self.blue_total_p_loss += p_loss
                else:
                    cell.fill = self.green_fill
                    self.green_total_p_loss += p_loss
            else:
                # ถ้าไม่ใช่เวลา 24.00 ให้ใช้เงื่อนไขของวันปัจจุบัน
                if self.is_weekend_or_holiday(dt):
                    cell.fill = self.blue_fill
                    self.blue_total_p_loss += p_loss
                else:
                    t = dt.time()
                    if time(9, 15) <= t <= time(22, 0):
                        cell.fill = self.red_fill
                        self.red_total_p_loss += p_loss
                    else:
                        cell.fill = self.green_fill
                        self.green_total_p_loss += p_loss
        
        # บันทึกไฟล์
        x = datetime.now()
        day = x.strftime('%d')
        month = x.strftime('%m')
        year = int(x.strftime('%y')) + 43
        hours = x.strftime('%H')
        minutes = x.strftime('%M')
        date_now_be = f"{day}_{month}_{year}_{hours}{minutes}"
        
        # สร้างโฟลเดอร์ output ถ้ายังไม่มี
        output_dir = os.path.join('D:', 'Python', 'voltage_loss', 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        output_path = f'D:\\Python\\voltage_loss\\output\\merged_data_{date_now_be}_{phase}.xlsx'
        output_wb.save(output_path)
        
        # แสดงผลลัพธ์ใน UI
        self.output_path_var.set(output_path)
        self.red_p_loss_var.set(f"{self.red_total_p_loss:.4f}")
        self.green_p_loss_var.set(f"{self.green_total_p_loss:.4f}")
        self.blue_p_loss_var.set(f"{self.blue_total_p_loss:.4f}")
        self.total_p_loss_var.set(f"{self.red_total_p_loss + self.green_total_p_loss + self.blue_total_p_loss:.4f}")
        
        # แสดงข้อความแจ้งเตือนว่าเสร็จสิ้นแล้ว
        messagebox.showinfo("Success", f"ดำเนินการเสร็จสิ้น\nบันทึกไฟล์ไว้ที่:\n{output_path}")
        
if __name__ == "__main__":
    root = tk.Tk()
    app = VoltageAnalyzerApp(root)
    root.mainloop()