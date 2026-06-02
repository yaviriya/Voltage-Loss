import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, time, date, timedelta
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import numpy as np
from sklearn.linear_model import LinearRegression

# ใช้ตำแหน่งของแอปเป็น base path เพื่อให้ทำงานได้ทุกเครื่อง
# - กรณี build เป็น .exe (PyInstaller) ใช้โฟลเดอร์ของไฟล์ .exe
# - กรณีรันเป็นสคริปต์ปกติ ใช้โฟลเดอร์ของไฟล์ .py
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def resource_path(rel_path):
    """หาพาธของไฟล์ทรัพยากร (เช่น ไอคอน) ให้ทำงานได้ทั้งตอนรัน .py และ .exe

    ตอน build แบบ PyInstaller --onefile ไฟล์ที่ฝังไว้จะถูกแตกไปที่ sys._MEIPASS
    """
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel_path)


class VoltageAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Voltage Loss Calculation")
        self.root.geometry("720x860")

        # ตั้งไอคอนหน้าต่างแอป (ถ้าหาไฟล์ไม่เจอก็ข้ามไป ไม่ให้แอป crash)
        try:
            self.root.iconbitmap(resource_path("wave.ico"))
        except Exception:
            pass

        # ตัวแปรเก็บพาธของไฟล์ (Voltage, Current, Power Factor, Normal Voltage, Normal Current)
        self.input_paths = [None, None, None, None, None]
        self.multiply_factor_var = tk.StringVar(value="30")  # CT Ratio เริ่มต้น 30

        # ตัวแปรสำหรับเก็บผลรวมของ P Loss แยกตามสี (Peak / Off-Peak / Holiday)
        self.red_total_p_loss = 0
        self.green_total_p_loss = 0
        self.blue_total_p_loss = 0

        # เตรียมสีที่จะใช้
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')   # สีแดง
        self.green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')  # สีเขียว
        self.blue_fill = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')   # สีฟ้า

        # วันหยุดราชการ
        self.holidays = {
            date(2025, 1, 1),
            date(2025, 2, 12),
            date(2025, 4, 7),
            date(2025, 4, 14),
            date(2025, 4, 15),
            date(2025, 4, 16),
            date(2025, 5, 1),
            date(2025, 5, 5),
            date(2025, 5, 9),
            date(2025, 5, 12),
            date(2025, 6, 2),
            date(2025, 6, 3),
            date(2025, 7, 10),
            date(2025, 7, 11),
            date(2025, 7, 28),
            date(2025, 8, 11),
            date(2025, 8, 12),
            date(2025, 10, 13),
            date(2025, 10, 23),
            date(2025, 12, 5),
            date(2025, 12, 10),
            date(2025, 12, 31),
            # ใส่วันอื่น ๆ ตามต้องการ
        }

        self.create_widgets()

    def create_widgets(self):
        """สร้าง UI elements"""
        # file_path_vars เก็บตาม index ของ input_paths
        # (0=Voltage, 1=Current, 2=Power Factor, 3=Normal Voltage, 4=Normal Current)
        self.file_path_vars = [tk.StringVar() for _ in range(5)]

        def build_file_rows(parent, items):
            """สร้างแถวเลือกไฟล์: items = [(label, index), ...]"""
            for row, (label, idx) in enumerate(items):
                ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=5, pady=5)
                ttk.Entry(parent, textvariable=self.file_path_vars[idx], width=50).grid(row=row, column=1, padx=5, pady=5)
                ttk.Button(parent, text="Browse...", command=lambda i=idx: self.browse_file(i)).grid(row=row, column=2, padx=5, pady=5)

        # เฟรมไฟล์ Normal (ใช้ฝึกโมเดล regression) — อยู่ด้านบน
        normal_frame = ttk.LabelFrame(self.root, text="Normal Files")
        normal_frame.pack(fill="x", padx=10, pady=(10, 20))
        build_file_rows(normal_frame, [("Normal Voltage", 3), ("Normal Current", 4)])

        # เฟรมไฟล์ที่ใช้คำนวณ — อยู่ด้านล่าง (เว้นระยะห่างจากกลุ่ม Normal)
        measure_frame = ttk.LabelFrame(self.root, text="Calculation Files")
        measure_frame.pack(fill="x", padx=10, pady=(0, 10))
        build_file_rows(measure_frame, [("Voltage", 0), ("Current", 1), ("Power Factor", 2)])

        # สร้างเฟรมสำหรับตั้งค่าพารามิเตอร์ CT (Radio Button แนวนอน)
        param_frame = ttk.LabelFrame(self.root, text="CT")
        param_frame.pack(fill="x", padx=10, pady=10)

        # ตัวเลือก CT: (ป้ายที่แสดง, ค่า CT_Ratio)
        ct_options = [("100/5", "20"), ("150/5", "30"), ("250/5", "50"), ("400/5", "80")]
        for col, (label, value) in enumerate(ct_options):
            ttk.Radiobutton(
                param_frame, text=label, value=value, variable=self.multiply_factor_var
            ).grid(row=0, column=col, sticky="w", padx=15, pady=5)

        # สร้างปุ่มเริ่มการคำนวณ
        ttk.Button(self.root, text="Execute", command=self.process_files).pack(pady=20)

        # สร้างเฟรมสำหรับแสดงผลลัพธ์
        result_frame = ttk.LabelFrame(self.root, text="Results")
        result_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # สร้างตัวแปรสำหรับแสดงผลลัพธ์
        self.output_path_var = tk.StringVar()
        self.red_p_loss_var = tk.StringVar()
        self.green_p_loss_var = tk.StringVar()
        self.blue_p_loss_var = tk.StringVar()
        self.total_p_loss_var = tk.StringVar()
        self.r2_a_var = tk.StringVar()
        self.r2_b_var = tk.StringVar()
        self.r2_c_var = tk.StringVar()

        # แสดงพาธของไฟล์ผลลัพธ์
        ttk.Label(result_frame, text="Output File:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.output_path_var, width=70, state="readonly").grid(row=0, column=1, padx=5, pady=5)

        # แสดงค่า R-Square ของแต่ละโมเดล V_regression (ยิ่งใกล้ 1 ยิ่งดี)
        ttk.Label(result_frame, text="R² Phase A:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.r2_a_var, width=15, state="readonly").grid(row=1, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(result_frame, text="R² Phase B:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.r2_b_var, width=15, state="readonly").grid(row=2, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(result_frame, text="R² Phase C:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.r2_c_var, width=15, state="readonly").grid(row=3, column=1, sticky="w", padx=5, pady=5)

        # แสดงผลรวมของ P Loss แยกตามสี
        ttk.Label(result_frame, text="P Loss Peak:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.red_p_loss_var, width=15, state="readonly").grid(row=4, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=4, column=1, sticky="w", padx=(120, 5), pady=5)

        ttk.Label(result_frame, text="P Loss Off-Peak:").grid(row=5, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.green_p_loss_var, width=15, state="readonly").grid(row=5, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=5, column=1, sticky="w", padx=(120, 5), pady=5)

        ttk.Label(result_frame, text="P Loss Holiday:").grid(row=6, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.blue_p_loss_var, width=15, state="readonly").grid(row=6, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=6, column=1, sticky="w", padx=(120, 5), pady=5)

        ttk.Label(result_frame, text="Total P Loss:").grid(row=7, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(result_frame, textvariable=self.total_p_loss_var, width=15, state="readonly").grid(row=7, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(result_frame, text="kWh").grid(row=7, column=1, sticky="w", padx=(120, 5), pady=5)

        # สร้างปุ่มออกจากโปรแกรม
        ttk.Button(self.root, text="Exit", command=self.root.destroy).pack(pady=10)

    def browse_file(self, index):
        """เปิดหน้าต่างให้เลือกไฟล์ Excel"""
        file_path = filedialog.askopenfilename(
            title=f"Select Excel File {index+1}",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_paths[index] = file_path
            self.file_path_vars[index].set(file_path)

    def is_weekend_or_holiday(self, dt):
        """ตรวจสอบว่าเป็นวันเสาร์-อาทิตย์ หรือวันหยุดราชการ"""
        if dt.date() in self.holidays:
            return True
        if dt.weekday() >= 5:  # 5=Saturday, 6=Sunday
            return True
        return False

    def find_header_row(self, rows):
        """หาแถวที่เป็นหัวตาราง โดยหาแถวแรกที่คอลัมน์ A มีรูปแบบวันที่"""
        for row_idx, row in enumerate(rows, start=1):
            first = row[0] if row else None
            cell_a_value = str(first).strip() if first is not None else ""
            if self.is_valid_datetime(cell_a_value):
                if row_idx > 1:
                    return row_idx - 1
                else:
                    return row_idx
        return None

    def is_valid_datetime(self, text):
        """ตรวจสอบว่าข้อความเป็นวันที่หรือไม่"""
        if not text:
            return False
        return any(c.isdigit() for c in text) and ('/' in text or '-' in text)

    def load_rows(self, file_path):
        """โหลดทุกแถวจากไฟล์ Excel เป็น list ของ list (ค่าของแต่ละเซลล์)

        เลือก engine ตามนามสกุลไฟล์:
        - .xls            → xlrd (รูปแบบเก่า)
        - .xlsx / .xlsm   → openpyxl (รูปแบบใหม่)
        เซลล์ชนิดวันที่ใน .xls จะถูกแปลงเป็นข้อความ dd/mm/yyyy HH.MM
        เพื่อให้ตัวอ่านวันที่ (ที่ออกแบบไว้สำหรับข้อความ) ใช้งานได้
        """
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
                        row.append(dt.strftime("%d/%m/%Y %H.%M"))
                    else:
                        row.append(cell.value)
                rows.append(row)
            return rows

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]

    def read_excel_data(self, file_path):
        """อ่านข้อมูลจากไฟล์ Excel (รองรับทั้ง .xlsx และ .xls)"""
        try:
            rows = self.load_rows(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"เปิดไฟล์ไม่ได้: {os.path.basename(file_path)}\n{e}")
            return None
        header_row = self.find_header_row(rows)

        if header_row is None:
            messagebox.showwarning("Warning", f"ไม่พบแถวหัวตารางในไฟล์ {os.path.basename(file_path)}")
            return None

        data = {}
        is_thai_date = False
        # ข้อมูลเริ่มที่แถวถัดจากหัวตาราง (header_row เป็นเลขแถวแบบ 1-based)
        data_rows = rows[header_row:]

        # ตรวจสอบแถวแรกที่มีข้อมูลวันที่เพื่อดูว่าเป็นวันที่ไทยหรือไม่
        for row in data_rows:
            raw_value = str(row[0]) if row and row[0] is not None else ""
            if self.is_valid_datetime(raw_value):
                clean_value = raw_value.replace('\xa0', '').strip()
                if ' ' in clean_value:
                    parts = clean_value.split()
                    date_str = parts[0]
                else:
                    date_str = clean_value

                if '/' in date_str:
                    day, month, year = date_str.split('/')
                elif '-' in date_str:
                    day, month, year = date_str.split('-')
                else:
                    continue

                try:
                    year_int = int(year)
                    if year_int - 543 > 2000:
                        is_thai_date = True
                except ValueError:
                    pass
                break

        # อ่านข้อมูลทั้งหมด
        for row in data_rows:
            try:
                raw_value = str(row[0]) if row and row[0] is not None else ""
                if not self.is_valid_datetime(raw_value):
                    continue

                clean_value = raw_value.replace('\xa0', '').strip()

                if ' ' in clean_value:
                    parts = clean_value.split()
                    date_str = parts[0]
                    time_str = parts[1]
                else:
                    if '/' in clean_value:
                        date_parts = clean_value.split('/')
                        if len(date_parts) >= 3:
                            date_str = '/'.join(date_parts[:3])
                            time_str = "00:00"
                        else:
                            continue
                    elif '-' in clean_value:
                        date_parts = clean_value.split('-')
                        if len(date_parts) >= 3:
                            date_str = '-'.join(date_parts[:3])
                            time_str = "00:00"
                        else:
                            continue
                    else:
                        continue

                if '/' in date_str:
                    day, month, year = date_str.split('/')
                elif '-' in date_str:
                    day, month, year = date_str.split('-')
                else:
                    continue

                if is_thai_date:
                    year = str(int(year) - 543)

                std_date_str = f"{day}/{month}/{year}"

                if ':' in time_str:
                    time_parts = time_str.split(':')
                    hour = time_parts[0]
                    minute = time_parts[1]
                    if len(minute) > 2:
                        minute = minute[:2]
                    std_time_str = f"{hour}.{minute}"
                else:
                    std_time_str = time_str

                if std_time_str == "24.00" or std_time_str == "24:00" or (hour == "24" if 'hour' in locals() else False):
                    dt_date = datetime.strptime(std_date_str, "%d/%m/%Y")
                    dt = dt_date + timedelta(days=1)
                else:
                    std_date_time_str = f"{std_date_str} {std_time_str}"
                    try:
                        dt = datetime.strptime(std_date_time_str, "%d/%m/%Y %H.%M")
                    except ValueError:
                        if '.' in std_time_str:
                            h, m = std_time_str.split('.')
                            std_time_str = f"{h}:{m}"
                        elif ':' in std_time_str:
                            h, m = std_time_str.split(':')
                            std_time_str = f"{h}.{m}"
                        std_date_time_str = f"{std_date_str} {std_time_str}"
                        try:
                            dt = datetime.strptime(std_date_time_str, "%d/%m/%Y %H.%M")
                        except ValueError:
                            dt = datetime.strptime(std_date_str, "%d/%m/%Y")

                data[dt] = list(row)
            except Exception:
                continue

        return data, header_row, None

    def compute_v_regression(self, normal_voltage_data, normal_current_data):
        """ฝึกโมเดล Linear Regression 3 ตัว (เฟส A, B, C) จากข้อมูลช่วงปกติ

        จับคู่แถวจากไฟล์ Normal Voltage และ Normal Current ตาม datetime ที่ตรงกัน
        โดยสมมติว่าคอลัมน์ลำดับ 1, 2, 3 ของแต่ละไฟล์คือเฟส A, B, C ตามลำดับ
        ดังนั้นแต่ละแถวข้อมูลฝึกคือ [V_a, V_b, V_c, I_a, I_b, I_c]

        ตัวแปรต้น (X) ของแต่ละเฟส:
            A: [V_b, V_c, I_a, I_b, I_c]   (ทำนาย V_a)
            B: [V_a, V_c, I_a, I_b, I_c]   (ทำนาย V_b)
            C: [V_a, V_b, I_a, I_b, I_c]   (ทำนาย V_c)

        คืนค่า (models, r2s):
            models[ph] = LinearRegression ที่ฝึกแล้ว (ph เป็น 'A'/'B'/'C')
            r2s[ph]    = ค่า R-Square ของโมเดลนั้น
        ถ้าข้อมูลไม่พอจะคืน (None, None)
        """
        rows = []
        for dt in sorted(set(normal_voltage_data) & set(normal_current_data)):
            v = normal_voltage_data[dt]
            c = normal_current_data[dt]
            try:
                v_a, v_b, v_c = float(v[1]), float(v[2]), float(v[3])
                i_a, i_b, i_c = float(c[1]), float(c[2]), float(c[3])
            except (ValueError, TypeError, IndexError):
                continue
            rows.append((v_a, v_b, v_c, i_a, i_b, i_c))

        # ต้องมีจำนวนแถวมากกว่าจำนวนตัวแปรต้น (5) จึง regression ได้อย่างมีความหมาย
        if len(rows) <= 5:
            return None, None

        arr = np.array(rows, dtype=float)
        v_a, v_b, v_c = arr[:, 0], arr[:, 1], arr[:, 2]
        i_a, i_b, i_c = arr[:, 3], arr[:, 4], arr[:, 5]

        feature_map = {
            'A': np.column_stack([v_b, v_c, i_a, i_b, i_c]),
            'B': np.column_stack([v_a, v_c, i_a, i_b, i_c]),
            'C': np.column_stack([v_a, v_b, i_a, i_b, i_c]),
        }
        target_map = {'A': v_a, 'B': v_b, 'C': v_c}

        models, r2s = {}, {}
        for ph in ('A', 'B', 'C'):
            model = LinearRegression()
            model.fit(feature_map[ph], target_map[ph])
            models[ph] = model
            r2s[ph] = model.score(feature_map[ph], target_map[ph])
        return models, r2s

    def predict_v_regression(self, model, features):
        """ทำนายค่า V_regression ของแถวหนึ่งจากโมเดลที่ฝึกไว้

        features ต้องเรียงลำดับเหมือนตอนฝึก ถ้ามีค่าใดเป็น None จะคืน None
        """
        if model is None or any(f is None for f in features):
            return None
        return float(model.intercept_ + np.dot(model.coef_, features))

    def process_files(self):
        """ประมวลผลไฟล์ Excel ทั้งหมด"""
        # ตรวจสอบว่าได้เลือกไฟล์ครบทั้ง 5 ไฟล์หรือไม่
        if None in self.input_paths:
            messagebox.showerror("Error", "โปรดเลือกให้ครบทั้ง 5 ไฟล์เพื่อทำการคำนวณ")
            return

        try:
            self.multiply_factor = float(self.multiply_factor_var.get())
        except ValueError:
            messagebox.showerror("Error", "กรุณาป้อนตัวคูณเป็นตัวเลข")
            return

        # รีเซ็ตตัวแปรผลรวม P Loss
        self.red_total_p_loss = 0
        self.green_total_p_loss = 0
        self.blue_total_p_loss = 0

        # อ่านข้อมูลจาก 3 ไฟล์หลัก (Voltage, Current, Power Factor)
        all_data = []
        for path in self.input_paths[:3]:
            result = self.read_excel_data(path)
            if result is None:
                messagebox.showerror("Error", "บางไฟล์ไม่สามารถอ่านได้ กรุณาตรวจสอบไฟล์ของคุณ")
                return
            data, _, _ = result
            all_data.append(data)

        # อ่านไฟล์ Normal Voltage และ Normal Current เพื่อฝึกโมเดล regression
        normal_volt_result = self.read_excel_data(self.input_paths[3])
        if normal_volt_result is None:
            messagebox.showerror("Error", "ไม่สามารถอ่านไฟล์ Normal Voltage ได้")
            return
        normal_curr_result = self.read_excel_data(self.input_paths[4])
        if normal_curr_result is None:
            messagebox.showerror("Error", "ไม่สามารถอ่านไฟล์ Normal Current ได้")
            return
        normal_volt_data, _, _ = normal_volt_result
        normal_curr_data, _, _ = normal_curr_result

        models, r2s = self.compute_v_regression(normal_volt_data, normal_curr_data)
        if models is None:
            messagebox.showerror("Error", "ข้อมูลช่วงปกติ (Normal Voltage/Current) ไม่พอสำหรับฝึกโมเดล regression")
            return

        # รวมข้อมูลจากทั้ง 3 ไฟล์หลัก
        merged_data = defaultdict(list)
        for dt in sorted(set().union(*[d.keys() for d in all_data])):
            row_data = []
            if dt.hour == 0 and dt.minute == 0:
                prev_day = dt - timedelta(days=1)
                datetime_str = prev_day.strftime("%d/%m/%Y 24.00")
            else:
                datetime_str = dt.strftime("%d/%m/%Y %H.%M")
            row_data.append(datetime_str)

            for data in all_data:
                if dt in data:
                    row_data.extend([val for val in data[dt][1:] if val is not None])

            merged_data[dt] = row_data

        # สร้างไฟล์ Excel ใหม่
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active

        headers = [
            "DateTime",
            "V Phase A", "V Phase B", "V Phase C",
            "I Phase A", "I Phase B", "I Phase C",
            "Power Factor",
            "V Reg A", "V Reg B", "V Reg C",
            "V Loss A", "V Loss B", "V Loss C",
            "P Loss A", "P Loss B", "P Loss C",
            "P Loss Total",
        ]
        output_ws.append(headers)

        # ตำแหน่งคอลัมน์ที่ต้องใช้ (1-based)
        v_a_col = headers.index("V Phase A") + 1
        v_b_col = headers.index("V Phase B") + 1
        v_c_col = headers.index("V Phase C") + 1
        i_a_col = headers.index("I Phase A") + 1
        i_b_col = headers.index("I Phase B") + 1
        i_c_col = headers.index("I Phase C") + 1
        pf_col = headers.index("Power Factor") + 1
        p_loss_total_col = headers.index("P Loss Total") + 1

        max_input_index = max(v_a_col, v_b_col, v_c_col, i_a_col, i_b_col, i_c_col, pf_col) - 1

        for dt, row_data in sorted(merged_data.items()):
            # ดึงค่าจากแถว (None ถ้าไม่ครบ)
            def get_float(idx):
                if len(row_data) > idx and row_data[idx] is not None:
                    try:
                        return float(row_data[idx])
                    except (ValueError, TypeError):
                        return None
                return None

            v_a = get_float(v_a_col - 1)
            v_b = get_float(v_b_col - 1)
            v_c = get_float(v_c_col - 1)
            i_a = get_float(i_a_col - 1)
            i_b = get_float(i_b_col - 1)
            i_c = get_float(i_c_col - 1)
            pf = get_float(pf_col - 1)

            # ทำนาย V_regression ของแต่ละเฟสจากตัวแปรต้นของแถวนี้
            # (ลำดับ feature ต้องตรงกับตอนฝึกใน compute_v_regression)
            v_reg_a = self.predict_v_regression(models['A'], [v_b, v_c, i_a, i_b, i_c])
            v_reg_b = self.predict_v_regression(models['B'], [v_a, v_c, i_a, i_b, i_c])
            v_reg_c = self.predict_v_regression(models['C'], [v_a, v_b, i_a, i_b, i_c])

            # V Loss ต่อเฟส: ถ้า V < V_reg*0.975 → V_reg*0.975 - V, ไม่งั้น 0
            def calc_v_loss(v, v_reg):
                if v is None or v_reg is None:
                    return 0.0
                threshold = v_reg * 0.975
                return threshold - v if v < threshold else 0.0

            v_loss_a = calc_v_loss(v_a, v_reg_a)
            v_loss_b = calc_v_loss(v_b, v_reg_b)
            v_loss_c = calc_v_loss(v_c, v_reg_c)

            # P Loss ต่อเฟส: V_loss * I * PF * CT_Ratio / 4000
            def calc_p_loss(v_loss, i_val):
                if i_val is None or pf is None:
                    return 0.0
                return (v_loss * i_val * pf * self.multiply_factor) / 4000

            p_loss_a = calc_p_loss(v_loss_a, i_a)
            p_loss_b = calc_p_loss(v_loss_b, i_b)
            p_loss_c = calc_p_loss(v_loss_c, i_c)
            p_loss_total = p_loss_a + p_loss_b + p_loss_c

            # ขยาย row_data ให้ครบ schema (ถ้าข้อมูลขาด ให้เติม None)
            while len(row_data) <= max_input_index:
                row_data.append(None)

            # ต่อท้าย: V Reg A/B/C, V Loss A/B/C, P Loss A/B/C, P Loss Total
            row_data.extend([
                v_reg_a, v_reg_b, v_reg_c,
                v_loss_a, v_loss_b, v_loss_c,
                p_loss_a, p_loss_b, p_loss_c,
                p_loss_total,
            ])

            output_ws.append(row_data)

            # ใส่สีในคอลัมน์ P Loss Total ตามช่วงเวลา (Peak / Off-Peak / Holiday)
            cell = output_ws.cell(row=output_ws.max_row, column=p_loss_total_col)
            is_24_00 = dt.hour == 0 and dt.minute == 0

            if is_24_00:
                prev_day = dt - timedelta(days=1)
                if self.is_weekend_or_holiday(prev_day):
                    cell.fill = self.blue_fill
                    self.blue_total_p_loss += p_loss_total
                else:
                    cell.fill = self.green_fill
                    self.green_total_p_loss += p_loss_total
            else:
                if self.is_weekend_or_holiday(dt):
                    cell.fill = self.blue_fill
                    self.blue_total_p_loss += p_loss_total
                else:
                    t = dt.time()
                    if time(9, 15) <= t <= time(22, 0):
                        cell.fill = self.red_fill
                        self.red_total_p_loss += p_loss_total
                    else:
                        cell.fill = self.green_fill
                        self.green_total_p_loss += p_loss_total

        # บันทึกไฟล์ — ใช้ path สัมพัทธ์กับสคริปต์ เพื่อให้รันได้ทุกเครื่อง
        x = datetime.now()
        day = x.strftime('%d')
        month = x.strftime('%m')
        year = int(x.strftime('%y')) + 43
        hours = x.strftime('%H')
        minutes = x.strftime('%M')
        date_now_be = f"{day}_{month}_{year}_{hours}{minutes}"

        output_dir = os.path.join(SCRIPT_DIR, 'files', 'output')
        os.makedirs(output_dir, exist_ok=True)

        output_path = os.path.join(output_dir, f'volt_output_data_{date_now_be}.xlsx')
        try:
            output_wb.save(output_path)
        except PermissionError:
            messagebox.showerror("Error", f"ไม่สามารถบันทึกไฟล์ได้ — โปรดปิดไฟล์ที่เปิดอยู่แล้วลองอีกครั้ง:\n{output_path}")
            return

        # แสดงผลลัพธ์ใน UI
        self.output_path_var.set(output_path)
        self.r2_a_var.set(f"{r2s['A']:.4f}")
        self.r2_b_var.set(f"{r2s['B']:.4f}")
        self.r2_c_var.set(f"{r2s['C']:.4f}")
        self.red_p_loss_var.set(f"{self.red_total_p_loss:.4f}")
        self.green_p_loss_var.set(f"{self.green_total_p_loss:.4f}")
        self.blue_p_loss_var.set(f"{self.blue_total_p_loss:.4f}")
        self.total_p_loss_var.set(f"{self.red_total_p_loss + self.green_total_p_loss + self.blue_total_p_loss:.4f}")

        messagebox.showinfo("Success", f"ดำเนินการเสร็จสิ้น\nบันทึกไฟล์ไว้ที่:\n{output_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = VoltageAnalyzerApp(root)
    root.mainloop()
